"""
AIC2026 Multimodal Retrieval Evaluation Pipeline
=================================================
Reproducible, audit-grade evaluation pipeline for benchmark: `aic2026_team_eval_dev_v1`

Enforces benchmark contract:
1. Strict separation of DEV_L21_150 and DEV_CROSS_60 (no unweighted 210-query combined score).
2. Automated integrity check for query/GT manifests, query_id alignment, and frame_coordinate_semantics ("original_frame_idx").
3. Standardized scoring for KIS, QA, and TRAKE tasks using canonical metrics: R@1, R@5, R@20, R@50, R@100.
4. Multi-dimensional breakdown analysis (Difficulty, Task-tags, Video-level, Latency profiling).
5. Comprehensive reporting package: REPORT.md, multi-sheet Excel (TEAM_EVAL_v1_RESULTS.xlsx), JSON/CSV diagnostics.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import os
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Ensure project root is on sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

CODEBASE_ROOT = PROJECT_ROOT.parent / "CodeBase"
if CODEBASE_ROOT.exists() and str(CODEBASE_ROOT) not in sys.path:
    sys.path.insert(0, str(CODEBASE_ROOT))

# Enforce cache storage strictly on E: drive
os.environ["HF_HUB_OFFLINE"] = "1"
os.environ["TRANSFORMERS_OFFLINE"] = "1"
os.environ["HF_HOME"] = str(PROJECT_ROOT / ".cache" / "huggingface")
os.environ["TORCH_HOME"] = str(PROJECT_ROOT / ".cache" / "torch")
os.environ["TRANSFORMERS_CACHE"] = str(PROJECT_ROOT / ".cache" / "huggingface" / "hub")
os.environ["PIP_CACHE_DIR"] = str(PROJECT_ROOT / ".cache" / "pip")

from backend.retrieval_service import RetrievalService

logger = logging.getLogger("aic.team_eval")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")

RECALL_KS = [1, 5, 20, 50, 100]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Comprehensive reproducible evaluation pipeline for aic2026_team_eval_dev_v1"
    )
    parser.add_argument(
        "--bundle-dir",
        type=Path,
        default=PROJECT_ROOT / "data" / "aic2026_team_eval_dev_v1",
        help="Path to unzipped benchmark bundle directory.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "outputs" / "evaluation" / "team_eval_v1",
        help="Directory for evaluation artifacts, Excel sheet, and REPORT.md.",
    )
    parser.add_argument(
        "--top-k",
        type=int,
        default=100,
        help="Top-K predictions to evaluate per query (default: 100).",
    )
    parser.add_argument(
        "--fusion-mode",
        choices=("static", "dynamic", "manual"),
        default="static",
        help="RetrievalService fusion mode.",
    )
    parser.add_argument(
        "--weights",
        default=None,
        help="Manual weights as visual,ocr,asr,object for manual mode.",
    )
    parser.add_argument(
        "--dedup-window-seconds",
        type=float,
        default=4.0,
        help="Temporal dedup window passed to RetrievalService.",
    )
    parser.add_argument(
        "--margin-seconds",
        type=float,
        default=3.0,
        help="Temporal tolerance window (+/- seconds) for ground-truth frame localization (default: 3.0s).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional query limit per benchmark set for fast smoke testing.",
    )
    return parser.parse_args()


# ==============================================================================
# 1. BENCHMARK INTEGRITY VERIFICATION
# ==============================================================================

class BenchmarkIntegrityChecker:
    """Automated integrity validation for query/GT manifests to prevent data leakage and ensure contract compliance."""

    def __init__(self, bundle_dir: Path):
        self.bundle_dir = bundle_dir
        self.registry_path = bundle_dir / "benchmark_registry.json"

    def run_all_checks(self) -> Dict[str, Any]:
        logger.info("--- Starting Benchmark Integrity Check ---")
        if not self.registry_path.exists():
            raise FileNotFoundError(f"Missing benchmark_registry.json at {self.registry_path}")

        with open(self.registry_path, "r", encoding="utf-8") as f:
            registry = json.load(f)

        checks_results = {
            "registry_version": registry.get("registry_version"),
            "combined_unweighted_dev_score_allowed": registry.get("combined_unweighted_dev_score_allowed", False),
            "benchmarks": {},
        }

        if checks_results["combined_unweighted_dev_score_allowed"]:
            raise ValueError("CONTRACT VIOLATION: combined_unweighted_dev_score_allowed MUST be false!")

        benchmarks_dir = self.bundle_dir / "benchmarks"
        for bm_entry in registry.get("benchmarks", []):
            bm_id = bm_entry["benchmark_id"]
            role = bm_entry["role"]
            content_present = bm_entry.get("content_in_development_bundle", True)

            if not content_present or bm_id == "SEALED_FINAL_30":
                logger.info(f"Skipping held-out/sealed benchmark: {bm_id}")
                continue

            bm_path = benchmarks_dir / bm_id.lower()
            if not bm_path.exists():
                raise FileNotFoundError(f"Benchmark folder missing: {bm_path}")

            bm_report = self._check_single_benchmark(bm_id, bm_path)
            checks_results["benchmarks"][bm_id] = bm_report
            logger.info(f"Benchmark {bm_id} integrity check: PASSED (Queries: {bm_report['query_count']})")

        logger.info("--- Benchmark Integrity Verification: ALL CHECKS PASSED ---")
        return checks_results

    def _check_single_benchmark(self, bm_id: str, bm_path: Path) -> Dict[str, Any]:
        manifest_p = bm_path / "manifest.json"
        queries_p = bm_path / "queries.jsonl"
        gt_p = bm_path / "gt.jsonl"

        for p in (manifest_p, queries_p, gt_p):
            if not p.exists():
                raise FileNotFoundError(f"Missing benchmark file: {p}")

        with open(manifest_p, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        # 1. Read queries and GT
        queries = [json.loads(line) for line in queries_p.read_text(encoding="utf-8").strip().splitlines() if line]
        gts = [json.loads(line) for line in gt_p.read_text(encoding="utf-8").strip().splitlines() if line]

        # 2. Count match check
        manifest_count = manifest.get("query_count", 0)
        if manifest_count != len(queries):
            raise ValueError(f"[{bm_id}] Manifest query_count ({manifest_count}) != actual queries count ({len(queries)})")
        if len(queries) != len(gts):
            raise ValueError(f"[{bm_id}] Queries count ({len(queries)}) != GT count ({len(gts)})")

        # 3. ID alignment check
        query_ids = [q["query_id"] for q in queries]
        gt_ids = [g["query_id"] for g in gts]

        if len(query_ids) != len(set(query_ids)):
            raise ValueError(f"[{bm_id}] Duplicate query_ids found in queries.jsonl")
        if query_ids != gt_ids:
            raise ValueError(f"[{bm_id}] Mismatch between query_id sequence in queries.jsonl and gt.jsonl")

        # 4. Task counts check
        expected_tasks = manifest.get("task_counts", {})
        actual_tasks = defaultdict(int)
        for q in queries:
            actual_tasks[q["task"]] += 1

        for task, expected_c in expected_tasks.items():
            if actual_tasks[task] != expected_c:
                raise ValueError(f"[{bm_id}] Task count mismatch for {task}: expected {expected_c}, got {actual_tasks[task]}")

        # 5. Coordinate semantics check
        semantics = manifest.get("frame_coordinate_semantics", "")
        if semantics and "original_frame_idx" not in semantics:
            raise ValueError(f"[{bm_id}] Invalid frame_coordinate_semantics: {semantics}. Must be 'original_frame_idx'")
        elif not semantics:
            semantics = "original_frame_idx / raw original video frame coordinate (canonical)"

        return {
            "status": "PASS",
            "query_count": len(queries),
            "task_counts": dict(actual_tasks),
            "semantics": semantics,
            "queries_file": str(queries_p),
            "gt_file": str(gt_p),
        }


# ==============================================================================
# 2. EVALUATION & SCORING ENGINE
# ==============================================================================

class BenchmarkEvaluator:
    """Evaluates retrieval service predictions against ground truth for KIS, QA, and TRAKE tasks."""

    def __init__(
        self,
        service: RetrievalService,
        top_k: int = 100,
        margin_seconds: float = 3.0,
        fusion_mode: str = "static",
        dedup_window_seconds: float = 4.0,
    ):
        self.service = service
        self.top_k = top_k
        self.margin_seconds = margin_seconds
        self.fusion_mode = fusion_mode
        self.dedup_window_seconds = dedup_window_seconds
        self.video_fps_cache: Dict[str, float] = {}

    def get_fps(self, video_id: str) -> float:
        if video_id in self.video_fps_cache:
            return self.video_fps_cache[video_id]

        fps = 25.0  # Default standard FPS
        # Look up in service video_to_keyframes_map or map CSV if available
        if video_id in self.service.video_to_keyframes_map:
            df_v = self.service.video_to_keyframes_map[video_id]
            if "fps" in df_v.columns:
                val = df_v["fps"].dropna()
                if not val.empty and val.iloc[0] > 0:
                    fps = float(val.iloc[0])

        self.video_fps_cache[video_id] = fps
        return fps

    def check_hit(
        self,
        task: str,
        cand_video: str,
        cand_frame: int,
        gt: Dict[str, Any],
        fps: float,
    ) -> bool:
        correct_video = gt.get("correct_video", "")
        if cand_video != correct_video:
            return False

        margin_frames = int(round(self.margin_seconds * fps))

        if task in ("KIS", "QA"):
            acceptable = gt.get("acceptable_intervals", [])
            for start, end in acceptable:
                if (start - margin_frames) <= cand_frame <= (end + margin_frames):
                    return True
            return False

        elif task == "TRAKE":
            event_intervals = gt.get("event_intervals", [])
            for start, end in event_intervals:
                if (start - margin_frames) <= cand_frame <= (end + margin_frames):
                    return True
            return False

        return False

    def evaluate_benchmark_set(
        self,
        bm_id: str,
        bm_path: Path,
        limit: Optional[int] = None,
    ) -> Dict[str, Any]:
        logger.info(f"--- Evaluating Benchmark Set: {bm_id} ---")
        queries_p = bm_path / "queries.jsonl"
        gt_p = bm_path / "gt.jsonl"

        queries = [json.loads(line) for line in queries_p.read_text(encoding="utf-8").strip().splitlines() if line]
        gts = [json.loads(line) for line in gt_p.read_text(encoding="utf-8").strip().splitlines() if line]

        if limit is not None and limit > 0:
            logger.info(f"Applying limit: evaluating first {limit} queries out of {len(queries)}")
            queries = queries[:limit]
            gts = gts[:limit]

        gt_map = {g["query_id"]: g for g in gts}
        per_query_results = []
        latencies_ms = []

        start_eval_t = time.time()

        for idx, q_item in enumerate(queries):
            qid = q_item["query_id"]
            task = q_item["task"]
            q_text = q_item["query"]
            diff = q_item.get("difficulty", "medium")
            tags = q_item.get("tags", [])
            gt = gt_map[qid]
            correct_video = gt.get("correct_video", "")

            # Search execution with latency profiling
            t0 = time.perf_counter()
            search_res = self.service.search(
                query=q_text,
                top_k=self.top_k,
                fusion_mode=self.fusion_mode,
                dedup_window_seconds=self.dedup_window_seconds,
            )
            lat_ms = (time.perf_counter() - t0) * 1000.0
            latencies_ms.append(lat_ms)

            predictions = search_res.get("results", [])
            fps = self.get_fps(correct_video)

            first_hit_rank: Optional[int] = None
            first_hit_video_match_rank: Optional[int] = None
            hits_at_k = {f"hit_at_{k}": False for k in RECALL_KS}

            matched_pred_info: Dict[str, Any] = {}

            for pred_idx, pred in enumerate(predictions, start=1):
                c_vid = str(pred.get("video_id", ""))
                c_frame = int(pred.get("frame_idx", pred.get("actual_frame_id", 0)))

                if c_vid == correct_video and first_hit_video_match_rank is None:
                    first_hit_video_match_rank = pred_idx

                is_hit = self.check_hit(task, c_vid, c_frame, gt, fps)

                if is_hit:
                    if first_hit_rank is None:
                        first_hit_rank = pred_idx
                        matched_pred_info = pred
                    for k in RECALL_KS:
                        if pred_idx <= k:
                            hits_at_k[f"hit_at_{k}"] = True

            # Calculate reciprocal rank
            mrr = 1.0 / first_hit_rank if first_hit_rank is not None else 0.0

            q_result = {
                "benchmark_id": bm_id,
                "query_id": qid,
                "task": task,
                "difficulty": diff,
                "tags": ",".join(tags) if isinstance(tags, list) else str(tags),
                "query": q_text,
                "correct_video": correct_video,
                "latency_ms": round(lat_ms, 2),
                "first_hit_rank": first_hit_rank if first_hit_rank is not None else "",
                "first_video_rank": first_hit_video_match_rank if first_hit_video_match_rank is not None else "",
                "mrr": round(mrr, 4),
                "top1_video": predictions[0].get("video_id", "") if predictions else "",
                "top1_frame": predictions[0].get("frame_idx", 0) if predictions else 0,
                "top1_score": round(float(predictions[0].get("score", 0.0)), 4) if predictions else 0.0,
            }
            for k in RECALL_KS:
                q_result[f"R@{k}"] = 1.0 if hits_at_k[f"hit_at_{k}"] else 0.0

            per_query_results.append(q_result)

            if (idx + 1) % 25 == 0 or (idx + 1) == len(queries):
                logger.info(f"[{bm_id}] Evaluated {idx+1}/{len(queries)} queries... Mean Latency: {np.mean(latencies_ms):.1f} ms")

        eval_duration = time.time() - start_eval_t
        df_queries = pd.DataFrame(per_query_results)

        # Aggregate Overall Metrics
        overall_metrics = self._aggregate_metrics(df_queries, latencies_ms)
        task_breakdown = self._group_metrics(df_queries, "task")
        diff_breakdown = self._group_metrics(df_queries, "difficulty")
        tag_breakdown = self._aggregate_tag_metrics(df_queries)
        video_breakdown = self._group_metrics(df_queries, "correct_video")

        return {
            "benchmark_id": bm_id,
            "num_queries": len(queries),
            "eval_duration_sec": round(eval_duration, 2),
            "overall": overall_metrics,
            "by_task": task_breakdown,
            "by_difficulty": diff_breakdown,
            "by_tag": tag_breakdown,
            "by_video": video_breakdown,
            "per_query": per_query_results,
        }

    def _aggregate_metrics(self, df: pd.DataFrame, latencies: List[float]) -> Dict[str, Any]:
        n = len(df)
        if n == 0:
            return {}

        metrics = {"num_queries": n}
        for k in RECALL_KS:
            metrics[f"R@{k}"] = round(float(df[f"R@{k}"].mean()), 4)

        metrics["MRR"] = round(float(df["mrr"].mean()), 4)

        # Latency statistics
        lats = np.array(latencies)
        metrics["latency"] = {
            "mean_ms": round(float(np.mean(lats)), 2),
            "median_ms": round(float(np.median(lats)), 2),
            "p95_ms": round(float(np.percentile(lats, 95)), 2),
            "p99_ms": round(float(np.percentile(lats, 99)), 2),
        }
        return metrics

    def _group_metrics(self, df: pd.DataFrame, group_col: str) -> Dict[str, Dict[str, Any]]:
        res = {}
        for key, group_df in df.groupby(group_col):
            lats = group_df["latency_ms"].tolist()
            res[str(key)] = self._aggregate_metrics(group_df, lats)
        return res

    def _aggregate_tag_metrics(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        tag_rows = defaultdict(list)
        for _, row in df.iterrows():
            tags_str = row.get("tags", "")
            tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            for t in tags:
                tag_rows[t].append(row)

        res = {}
        for tag, rows in sorted(tag_rows.items()):
            sub_df = pd.DataFrame(rows)
            res[tag] = self._aggregate_metrics(sub_df, sub_df["latency_ms"].tolist())
        return res


# ==============================================================================
# 3. EXCEL REPORT GENERATOR (TEAM_EVAL_v1_RESULTS.xlsx)
# ==============================================================================

class ExcelReportGenerator:
    """Generates styled multi-sheet Excel workbook TEAM_EVAL_v1_RESULTS.xlsx."""

    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        self.wb.remove(self.wb.active)

        # Styling definitions
        self.font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        self.font_section = Font(name="Calibri", size=13, bold=True, color="2E75B6")
        self.font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.font_bold = Font(name="Calibri", size=11, bold=True)
        self.font_normal = Font(name="Calibri", size=11)

        self.fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.fill_subheader = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        self.fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
        self.fill_highlight = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        self.fill_miss = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        self.thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

    def generate(self, results: Dict[str, Dict[str, Any]]) -> None:
        logger.info(f"Generating Excel workbook: {self.output_path}")

        self._build_summary_sheet(results)
        for bm_id in ("DEV_L21_150", "DEV_CROSS_60"):
            if bm_id in results:
                self._build_benchmark_metrics_sheet(bm_id, results[bm_id])

        self._build_breakdown_sheet("Breakdown_Difficulty", results, "by_difficulty")
        self._build_breakdown_sheet("Breakdown_Tags", results, "by_tag")
        self._build_per_query_sheet(results)
        self._build_failure_analysis_sheet(results)
        self._build_latency_profile_sheet(results)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(self.output_path))
        logger.info(f"Excel workbook saved successfully to {self.output_path}")

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value if cell.value is not None else "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

    def _build_summary_sheet(self, results: Dict[str, Dict[str, Any]]):
        ws = self.wb.create_sheet(title="Summary")
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value="AIC2026 Multimodal Retrieval Evaluation - Executive Summary").font = self.font_title
        ws.cell(row=2, column=1, value=f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Benchmark Contract: STRICT ISOLATION").font = Font(name="Calibri", size=10, italic=True, color="595959")

        # Overview Table Header
        headers = ["Benchmark Set", "Role", "Queries", "R@1", "R@5", "R@20", "R@50", "R@100", "MRR", "Mean Latency (ms)"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c_idx, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx = 5
        for bm_id, bm_data in results.items():
            ov = bm_data["overall"]
            role = "PUBLIC_REGRESSION_DEBUG" if bm_id == "DEV_L21_150" else "PUBLIC_CROSS_LEVEL"
            vals = [
                bm_id,
                role,
                bm_data["num_queries"],
                ov.get("R@1", 0),
                ov.get("R@5", 0),
                ov.get("R@20", 0),
                ov.get("R@50", 0),
                ov.get("R@100", 0),
                ov.get("MRR", 0),
                ov.get("latency", {}).get("mean_ms", 0),
            ]
            for c_idx, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_idx, column=c_idx, value=val)
                cell.font = self.font_bold if c_idx == 1 else self.font_normal
                cell.border = self.thin_border
                if 4 <= c_idx <= 9:
                    cell.number_format = "0.00%" if c_idx <= 8 else "0.0000"
                    cell.alignment = Alignment(horizontal="right")
                elif c_idx in (3, 10):
                    cell.alignment = Alignment(horizontal="center")
            row_idx += 1

        self._auto_fit_columns(ws)

    def _build_benchmark_metrics_sheet(self, bm_id: str, bm_data: Dict[str, Any]):
        ws = self.wb.create_sheet(title=f"{bm_id}_Metrics")
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value=f"Canonical Evaluation Metrics: {bm_id}").font = self.font_title

        # Overall Table
        ws.cell(row=3, column=1, value="1. Overall Performance").font = self.font_section
        headers = ["Metric", "Value"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header

        r_idx = 5
        ov = bm_data["overall"]
        for k in RECALL_KS:
            ws.cell(row=r_idx, column=1, value=f"Recall@{k} (R@{k})").font = self.font_bold
            c = ws.cell(row=r_idx, column=2, value=ov.get(f"R@{k}", 0))
            c.number_format = "0.00%"
            r_idx += 1

        ws.cell(row=r_idx, column=1, value="Mean Reciprocal Rank (MRR)").font = self.font_bold
        ws.cell(row=r_idx, column=2, value=ov.get("MRR", 0)).number_format = "0.0000"
        r_idx += 2

        # Task Breakdown Table
        ws.cell(row=r_idx, column=1, value="2. Performance Breakdown by Task (KIS / QA / TRAKE)").font = self.font_section
        r_idx += 1
        t_headers = ["Task", "Queries", "R@1", "R@5", "R@20", "R@50", "R@100", "MRR", "Mean Latency (ms)"]
        for c, h in enumerate(t_headers, start=1):
            cell = ws.cell(row=r_idx, column=c, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_subheader
        r_idx += 1

        for task, t_data in bm_data.get("by_task", {}).items():
            vals = [
                task,
                t_data.get("num_queries", 0),
                t_data.get("R@1", 0),
                t_data.get("R@5", 0),
                t_data.get("R@20", 0),
                t_data.get("R@50", 0),
                t_data.get("R@100", 0),
                t_data.get("MRR", 0),
                t_data.get("latency", {}).get("mean_ms", 0),
            ]
            for c, val in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.border = self.thin_border
                if 3 <= c <= 8:
                    cell.number_format = "0.00%" if c <= 7 else "0.0000"
            r_idx += 1

        self._auto_fit_columns(ws)

    def _build_breakdown_sheet(self, sheet_name: str, results: Dict[str, Dict[str, Any]], key_name: str):
        ws = self.wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value=f"Multi-Dimensional Analysis: {sheet_name.replace('_', ' ')}").font = self.font_title

        r_idx = 3
        for bm_id, bm_data in results.items():
            ws.cell(row=r_idx, column=1, value=f"Benchmark: {bm_id}").font = self.font_section
            r_idx += 1

            headers = ["Category", "Queries", "R@1", "R@5", "R@20", "R@50", "R@100", "MRR", "Mean Latency (ms)"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c, value=h)
                cell.font = self.font_header
                cell.fill = self.fill_header
            r_idx += 1

            sub_map = bm_data.get(key_name, {})
            for cat, c_data in sub_map.items():
                vals = [
                    cat,
                    c_data.get("num_queries", 0),
                    c_data.get("R@1", 0),
                    c_data.get("R@5", 0),
                    c_data.get("R@20", 0),
                    c_data.get("R@50", 0),
                    c_data.get("R@100", 0),
                    c_data.get("MRR", 0),
                    c_data.get("latency", {}).get("mean_ms", 0),
                ]
                for c, val in enumerate(vals, start=1):
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.border = self.thin_border
                    if 3 <= c <= 8:
                        cell.number_format = "0.00%" if c <= 7 else "0.0000"
                r_idx += 1
            r_idx += 2

        self._auto_fit_columns(ws)

    def _build_per_query_sheet(self, results: Dict[str, Dict[str, Any]]):
        ws = self.wb.create_sheet(title="Per_Query_Details")
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "Benchmark Set", "Query ID", "Task", "Difficulty", "Tags",
            "Query Text", "Target Video", "Hit Rank", "Video Match Rank",
            "MRR", "R@1", "R@5", "R@20", "R@50", "R@100", "Latency (ms)"
        ]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header

        r_idx = 2
        for bm_id, bm_data in results.items():
            for q in bm_data.get("per_query", []):
                vals = [
                    bm_id, q["query_id"], q["task"], q["difficulty"], q["tags"],
                    q["query"], q["correct_video"], q["first_hit_rank"], q["first_video_rank"],
                    q["mrr"], q["R@1"], q["R@5"], q["R@20"], q["R@50"], q["R@100"], q["latency_ms"]
                ]
                for c, val in enumerate(vals, start=1):
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.border = self.thin_border
                    if 11 <= c <= 15:
                        cell.number_format = "0%"
                        if val == 1:
                            cell.fill = self.fill_highlight
                    elif c == 8 and val == "":
                        cell.fill = self.fill_miss
                r_idx += 1

        self._auto_fit_columns(ws)

    def _build_failure_analysis_sheet(self, results: Dict[str, Dict[str, Any]]):
        ws = self.wb.create_sheet(title="Failure_Analysis")
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value="Diagnostic Failure Analysis & Best Performing Cases").font = self.font_title

        r_idx = 3
        headers = ["Benchmark Set", "Query ID", "Task", "Difficulty", "Tags", "Query Text", "Target Video", "Top-1 Pred Video", "Failure Diagnostic Category"]

        for bm_id, bm_data in results.items():
            ws.cell(row=r_idx, column=1, value=f"Zero-Hit Queries (Misses in Top-100) - {bm_id}").font = self.font_section
            r_idx += 1
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c, value=h)
                cell.font = self.font_header
                cell.fill = self.fill_subheader
            r_idx += 1

            misses = [q for q in bm_data.get("per_query", []) if q["first_hit_rank"] == ""]
            for q in misses:
                diag = "Wrong Video Retrieved" if q["first_video_rank"] == "" else f"Correct Video at Rank #{q['first_video_rank']} (Frame Misalignment)"
                vals = [bm_id, q["query_id"], q["task"], q["difficulty"], q["tags"], q["query"], q["correct_video"], q["top1_video"], diag]
                for c, val in enumerate(vals, start=1):
                    cell = ws.cell(row=r_idx, column=c, value=val)
                    cell.border = self.thin_border
                    if c == 9:
                        cell.fill = self.fill_miss
                r_idx += 1
            r_idx += 2

        self._auto_fit_columns(ws)

    def _build_latency_profile_sheet(self, results: Dict[str, Dict[str, Any]]):
        ws = self.wb.create_sheet(title="Latency_Profile")
        ws.views.sheetView[0].showGridLines = True

        ws.cell(row=1, column=1, value="System Latency & Throughput Profile").font = self.font_title

        headers = ["Benchmark Set", "Queries Evaluated", "Total Duration (s)", "Mean Latency (ms)", "Median Latency (ms)", "P95 Latency (ms)", "P99 Latency (ms)", "Throughput (QPS)"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_header

        r_idx = 4
        for bm_id, bm_data in results.items():
            ov_lat = bm_data["overall"].get("latency", {})
            dur = bm_data.get("eval_duration_sec", 1.0)
            n_q = bm_data.get("num_queries", 1)
            qps = round(n_q / max(dur, 0.001), 2)

            vals = [
                bm_id, n_q, dur, ov_lat.get("mean_ms", 0), ov_lat.get("median_ms", 0),
                ov_lat.get("p95_ms", 0), ov_lat.get("p99_ms", 0), qps
            ]
            for c, val in enumerate(vals, start=1):
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.border = self.thin_border
            r_idx += 1

        self._auto_fit_columns(ws)


# ==============================================================================
# 4. REPORT.md GENERATOR
# ==============================================================================

class MarkdownReportGenerator:
    """Generates research-grade REPORT.md evaluation report."""

    def __init__(self, output_path: Path):
        self.output_path = output_path

    def generate(self, results: Dict[str, Dict[str, Any]], integrity_report: Dict[str, Any]) -> None:
        logger.info(f"Generating research report: {self.output_path}")

        md = []
        md.append("# AIC2026 Multimodal Retrieval Evaluation Report (`aic2026_team_eval_dev_v1`)\n")
        md.append(f"**Generated Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        md.append("**Evaluation Pipeline:** PTK-SGU Automated Multimodal Evaluator (v1.0)\n")
        md.append("**Benchmark Contract Status:** `PASSED` (Strict Separation enforced, Zero Aggregate Leakage)\n")
        md.append("\n---\n")

        # Executive Summary
        md.append("## 1. Executive Summary\n")
        md.append("This evaluation package presents a comprehensive benchmark of the AIC2026 Multimodal Retrieval System ")
        md.append("on the standardized `aic2026_team_eval_dev_v1` benchmark suite. The evaluation enforces strict isolation ")
        md.append("between **DEV_L21_150** (public regression debug) and **DEV_CROSS_60** (public cross-level development) ")
        md.append("without unweighted aggregation, maintaining high research-grade integrity.\n\n")

        # Integrity Check Summary
        md.append("## 2. Integrity Verification & Benchmark Contract Compliance\n")
        md.append("| Benchmark Set | Query Count | Task Breakdown (KIS/QA/TRAKE) | Frame Semantics | Contract Status |\n")
        md.append("|---|:---:|:---:|:---:|:---:|\n")
        for bm_id, bm_info in integrity_report.get("benchmarks", {}).items():
            tb = bm_info.get("task_counts", {})
            tb_str = f"KIS:{tb.get('KIS',0)} / QA:{tb.get('QA',0)} / TRAKE:{tb.get('TRAKE',0)}"
            md.append(f"| `{bm_id}` | {bm_info['query_count']} | {tb_str} | `original_frame_idx` | **{bm_info['status']}** |\n")
        md.append("\n")

        # Overall Results Table
        md.append("## 3. Overall Retrieval Performance\n")
        md.append("Canonical metrics measured across Top-K predictions ($K \\in \\{1, 5, 20, 50, 100\\}$):\n\n")
        md.append("| Benchmark Set | Role | Queries | R@1 | R@5 | R@20 | R@50 | R@100 | MRR | Mean Latency |\n")
        md.append("|---|---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|:---:|\n")

        for bm_id, bm_data in results.items():
            ov = bm_data["overall"]
            role = "Regression Debug" if bm_id == "DEV_L21_150" else "Cross-Level Dev"
            lat = ov.get("latency", {}).get("mean_ms", 0)
            md.append(
                f"| **{bm_id}** | {role} | {bm_data['num_queries']} | "
                f"**{ov.get('R@1',0):.2%}** | **{ov.get('R@5',0):.2%}** | **{ov.get('R@20',0):.2%}** | "
                f"**{ov.get('R@50',0):.2%}** | **{ov.get('R@100',0):.2%}** | **{ov.get('MRR',0):.4f}** | {lat:.1f} ms |\n"
            )
        md.append("\n")

        # Task-level Breakdown
        md.append("## 4. Task-Level Performance Breakdown (KIS, QA, TRAKE)\n\n")
        for bm_id, bm_data in results.items():
            md.append(f"### Benchmark Set: `{bm_id}`\n")
            md.append("| Task | Queries | R@1 | R@5 | R@20 | R@50 | R@100 | MRR | Mean Latency |\n")
            md.append("|---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|:---:|\n")
            for task, t_data in bm_data.get("by_task", {}).items():
                lat = t_data.get("latency", {}).get("mean_ms", 0)
                md.append(
                    f"| **{task}** | {t_data.get('num_queries',0)} | "
                    f"{t_data.get('R@1',0):.2%} | {t_data.get('R@5',0):.2%} | {t_data.get('R@20',0):.2%} | "
                    f"{t_data.get('R@50',0):.2%} | {t_data.get('R@100',0):.2%} | {t_data.get('MRR',0):.4f} | {lat:.1f} ms |\n"
                )
            md.append("\n")

        # Multi-dimensional Breakdown (Difficulty & Tags)
        md.append("## 5. Multi-Dimensional Performance Breakdown\n\n")
        for bm_id, bm_data in results.items():
            md.append(f"### `{bm_id}` - Difficulty Breakdown\n")
            md.append("| Difficulty | Queries | R@1 | R@5 | R@20 | R@50 | R@100 | MRR |\n")
            md.append("|---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|\n")
            for diff, d_data in bm_data.get("by_difficulty", {}).items():
                md.append(
                    f"| **{diff.capitalize()}** | {d_data.get('num_queries',0)} | "
                    f"{d_data.get('R@1',0):.2%} | {d_data.get('R@5',0):.2%} | {d_data.get('R@20',0):.2%} | "
                    f"{d_data.get('R@50',0):.2%} | {d_data.get('R@100',0):.2%} | {d_data.get('MRR',0):.4f} |\n"
                )
            md.append("\n")

            md.append(f"### `{bm_id}` - Task Tag Breakdown\n")
            md.append("| Tag | Queries | R@1 | R@5 | R@20 | R@50 | R@100 | MRR |\n")
            md.append("|---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|\n")
            for tag, tg_data in bm_data.get("by_tag", {}).items():
                md.append(
                    f"| `{tag}` | {tg_data.get('num_queries',0)} | "
                    f"{tg_data.get('R@1',0):.2%} | {tg_data.get('R@5',0):.2%} | {tg_data.get('R@20',0):.2%} | "
                    f"{tg_data.get('R@50',0):.2%} | {tg_data.get('R@100',0):.2%} | {tg_data.get('MRR',0):.4f} |\n"
                )
            md.append("\n")

        # Latency Profile
        md.append("## 6. System Latency & Throughput Profile\n\n")
        md.append("| Benchmark Set | Mean Latency | Median (P50) | P95 Latency | P99 Latency | Throughput (QPS) |\n")
        md.append("|---|:---:|:---:|:---:|:---:|:---:|\n")
        for bm_id, bm_data in results.items():
            lats = bm_data["overall"].get("latency", {})
            dur = bm_data.get("eval_duration_sec", 1.0)
            n_q = bm_data.get("num_queries", 1)
            qps = n_q / max(dur, 0.001)
            md.append(f"| `{bm_id}` | {lats.get('mean_ms',0):.1f} ms | {lats.get('median_ms',0):.1f} ms | {lats.get('p95_ms',0):.1f} ms | {lats.get('p99_ms',0):.1f} ms | **{qps:.2f} QPS** |\n")
        md.append("\n")

        # Failure Analysis
        md.append("## 7. Failure Analysis & System Bottlenecks\n\n")
        for bm_id, bm_data in results.items():
            misses = [q for q in bm_data.get("per_query", []) if q["first_hit_rank"] == ""]
            hits1 = [q for q in bm_data.get("per_query", []) if q["first_hit_rank"] == 1]

            md.append(f"### `{bm_id}` Diagnostic Overview\n")
            md.append(f"- **Total Zero-Hit Queries (Misses @ K=100):** `{len(misses)}` / `{bm_data['num_queries']}` ({len(misses)/bm_data['num_queries']:.1%})\n")
            md.append(f"- **Total Top-1 Exact Hits:** `{len(hits1)}` / `{bm_data['num_queries']}` ({len(hits1)/bm_data['num_queries']:.1%})\n\n")

            if misses:
                md.append("#### Top Sample Missed Queries (Zero-Hit Diagnostic):\n")
                md.append("| Query ID | Task | Difficulty | Query Text | Target Video | Diagnostic Category |\n")
                md.append("|---|:---:|:---:|---|:---:|---|\n")
                for q in misses[:5]:
                    diag = "Video Selection Error" if q["first_video_rank"] == "" else f"Temporal Misalignment (Video @ Rank #{q['first_video_rank']})"
                    md.append(f"| `{q['query_id']}` | {q['task']} | {q['difficulty']} | {q['query']} | `{q['correct_video']}` | {diag} |\n")
                md.append("\n")

        # Actionable Optimization Recommendations
        md.append("## 8. Actionable Optimization Recommendations\n\n")
        md.append("1. **Temporal Alignment Enhancement:** Incorporate dynamic temporal window expansion for long-segment queries (TRAKE & multi-event KIS) to mitigate frame boundary mismatches.\n")
        md.append("2. **BGE Reranker Weight Tuning:** Enhance text-branch late fusion weights for high-density OCR/ASR queries to increase R@1 accuracy on speech/text-heavy samples.\n")
        md.append("3. **Multi-Event Graph TRAKE Search:** Integrate event graph structural search to preserve temporal ordering invariants during candidate ranking.\n")

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.output_path.write_text("".join(md), encoding="utf-8")
        logger.info(f"REPORT.md successfully written to {self.output_path}")


# ==============================================================================
# MAIN EXECUTION FLOW
# ==============================================================================

def main():
    args = parse_args()

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Run Automated Integrity Check
    checker = BenchmarkIntegrityChecker(args.bundle_dir)
    integrity_report = checker.run_all_checks()

    # Save integrity check report
    with open(output_dir / "integrity_report.json", "w", encoding="utf-8") as f:
        json.dump(integrity_report, f, indent=2)

    # 2. Initialize Singleton Retrieval Service
    logger.info("Initializing Singleton RetrievalService for online inference...")
    service = RetrievalService.get_instance()
    service.initialize()

    # 3. Instantiate Evaluator
    evaluator = BenchmarkEvaluator(
        service=service,
        top_k=args.top_k,
        margin_seconds=args.margin_seconds,
        fusion_mode=args.fusion_mode,
        dedup_window_seconds=args.dedup_window_seconds,
    )

    results_per_benchmark: Dict[str, Dict[str, Any]] = {}

    benchmarks_dir = args.bundle_dir / "benchmarks"
    for bm_id in ("DEV_L21_150", "DEV_CROSS_60"):
        bm_path = benchmarks_dir / bm_id.lower()
        if bm_path.exists():
            bm_res = evaluator.evaluate_benchmark_set(bm_id, bm_path, limit=args.limit)
            results_per_benchmark[bm_id] = bm_res

            # Save per-benchmark raw JSON & CSV
            bm_out_dir = output_dir / bm_id.lower()
            bm_out_dir.mkdir(parents=True, exist_ok=True)

            with open(bm_out_dir / "metrics_summary.json", "w", encoding="utf-8") as f:
                json.dump(bm_res, f, indent=2)

            df_per_query = pd.DataFrame(bm_res["per_query"])
            df_per_query.to_csv(bm_out_dir / "per_query_details.csv", index=False, encoding="utf-8-sig")

    # Save combined results summary JSON
    with open(output_dir / "combined_evaluation_results.json", "w", encoding="utf-8") as f:
        json.dump(results_per_benchmark, f, indent=2)

    # 4. Generate Excel Report (TEAM_EVAL_v1_RESULTS.xlsx)
    excel_gen = ExcelReportGenerator(output_dir / "TEAM_EVAL_v1_RESULTS.xlsx")
    excel_gen.generate(results_per_benchmark)

    # 5. Generate Markdown Report (REPORT.md)
    md_gen = MarkdownReportGenerator(output_dir / "REPORT.md")
    md_gen.generate(results_per_benchmark, integrity_report)

    logger.info("=================================================================")
    logger.info("EVALUATION PIPELINE COMPLETED SUCCESSFULLY!")
    logger.info(f"Artifacts generated in: {output_dir}")
    logger.info("=================================================================")


if __name__ == "__main__":
    main()
